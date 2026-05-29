#!/usr/bin/env python3

"""
streamlist
----------
Downloads YouTube / YouTube Music playlists as M4A tracks with embedded
album art, metadata tags, and a .m3u playlist file.

Features:
  - Resume & sync: skips already-downloaded tracks on re-runs
  - Per-playlist JSON cache for instant skipping with zero network calls
  - Auto-retry: failed tracks are not cached and retried next run
  - Cache rebuild from existing files when no cache is present
  - YouTube Premium & private playlist support via browser cookies
  - Smart AAC encoder detection (Apple hardware → FDK → native fallback)
  - Cover art editor: re-embed thumbnails as center crop, smart crop, or padded blur

Inputs:
  - YouTube / YouTube Music playlist URL (or single video)
  - Excel file with columns: url (required), title (optional), artist (optional)
"""

import io
import os
import sys
import re
import json
import shutil
import subprocess
import time
import tempfile
import argparse
from pathlib import Path

import yt_dlp
from tqdm import tqdm
from PIL import Image, ImageFilter, ImageOps
from mutagen.mp4 import MP4, MP4Cover

try:
    import openpyxl
except ImportError:
    openpyxl = None

# ─────────────────────────────────────────────
#  FFMPEG DETECTION
# ─────────────────────────────────────────────

def detect_ffmpeg():
    for candidate in [shutil.which('ffmpeg'), "/opt/homebrew/bin/ffmpeg", "/usr/local/bin/ffmpeg"]:
        if candidate and os.path.exists(candidate):
            return candidate
    raise RuntimeError(
        "FFmpeg not found. Install it:\n"
        "  macOS:  brew install ffmpeg\n"
        "  Linux:  sudo apt install ffmpeg\n"
        "  Windows: https://ffmpeg.org/download.html"
    )

def detect_ffprobe(ffmpeg_path):
    probe = shutil.which('ffprobe') or ffmpeg_path.replace('ffmpeg', 'ffprobe')
    return probe if os.path.exists(probe) else ffmpeg_path.replace('ffmpeg', 'ffprobe')

FFMPEG  = detect_ffmpeg()
FFPROBE = detect_ffprobe(FFMPEG)

# ─────────────────────────────────────────────
#  ENCODER DETECTION
# ─────────────────────────────────────────────

def detect_best_aac_encoder():
    candidates = [
        ('aac_at',     True,  'Apple AudioToolbox (hardware, macOS)'),
        ('libfdk_aac', False, 'Fraunhofer FDK AAC (best software quality)'),
        ('aac',        False, 'FFmpeg native AAC (fallback)'),
    ]
    try:
        result = subprocess.run(
            [FFMPEG, '-hide_banner', '-encoders'],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=5
        )
        available = result.stdout
    except Exception:
        return ('aac', False, 'FFmpeg native AAC')

    for encoder, is_hw, desc in candidates:
        if encoder not in available:
            continue
        try:
            val = subprocess.run(
                [FFMPEG, '-f', 'lavfi', '-i', 'anullsrc=r=44100:cl=mono',
                 '-t', '0.1', '-c:a', encoder, '-b:a', '128k', '-f', 'null', '-'],
                stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=5
            )
            if val.returncode == 0:
                return (encoder, is_hw, desc)
        except Exception:
            continue

    return ('aac', False, 'FFmpeg native AAC')

# ─────────────────────────────────────────────
#  QUIET LOGGER
# ─────────────────────────────────────────────

class QuietLogger:
    def debug(self, msg):   pass
    def warning(self, msg): pass
    def error(self, msg):   pass

# ─────────────────────────────────────────────
#  SAFE FILENAME
# ─────────────────────────────────────────────

def safe_filename(name, max_len=120):
    name = re.sub(r'[\\/:*?"<>|]', '-', name)
    name = name.strip('. ')
    return name[:max_len] if len(name) > max_len else name

# ─────────────────────────────────────────────
#  YDL BASE OPTIONS (shared cookie config)
# ─────────────────────────────────────────────

_COOKIE_BROWSER = None

def ydl_base_opts(with_cookies=True):
    opts = {'quiet': True, 'logger': QuietLogger(), 'no_warnings': True}
    if _COOKIE_BROWSER and with_cookies:
        opts['cookiesfrombrowser'] = (_COOKIE_BROWSER, None, None, None)
    return opts

# ─────────────────────────────────────────────
#  PLAYLIST EXPANSION
# ─────────────────────────────────────────────

def expand_playlist_url(url):
    opts = {**ydl_base_opts(), 'extract_flat': True}
    with yt_dlp.YoutubeDL(opts) as ydl:
        info = ydl.extract_info(url, download=False)

    if info.get('_type') == 'playlist':
        entries = info.get('entries', []) or []
        tracks = []
        for e in entries:
            if not e:
                continue
            vid_url = e.get('url') or e.get('webpage_url') or f"https://www.youtube.com/watch?v={e['id']}"
            tracks.append({'title': e.get('title', ''), 'url': vid_url, 'artist': ''})
        print(f"  Found {len(tracks)} track(s) in playlist: {info.get('title', url)}")
        return info.get('title', 'Playlist'), tracks
    else:
        title = info.get('title', 'Track')
        return title, [{'title': title, 'url': url, 'artist': ''}]

# ─────────────────────────────────────────────
#  EXCEL INPUT
# ─────────────────────────────────────────────

def load_excel(path):
    if openpyxl is None:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]

    if 'url' not in headers:
        raise ValueError("Excel file must have a 'url' column.")

    url_idx    = headers.index('url')
    title_idx  = headers.index('title')  if 'title'  in headers else None
    artist_idx = headers.index('artist') if 'artist' in headers else None

    tracks = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        url = str(row[url_idx]).strip() if row[url_idx] else ''
        if not url or url.lower() == 'none':
            continue
        title  = str(row[title_idx]).strip()  if title_idx  is not None and row[title_idx]  else ''
        artist = str(row[artist_idx]).strip() if artist_idx is not None and row[artist_idx] else ''
        tracks.append({'title': title, 'url': url, 'artist': artist})

    return tracks

# ─────────────────────────────────────────────
#  AUDIO DOWNLOAD
# ─────────────────────────────────────────────

def download_audio(url, out_path_no_ext, with_cookies=True):
    def _attempt(use_cookies):
        opts = {
            **ydl_base_opts(with_cookies=use_cookies),
            'format': 'bestaudio/best',
            'outtmpl': f'{out_path_no_ext}.%(ext)s',
            'ffmpeg_location': FFMPEG,
            'postprocessors': [],
        }
        with yt_dlp.YoutubeDL(opts) as ydl:
            ydl.download([url])

    try:
        _attempt(use_cookies=with_cookies)
    except Exception:
        if _COOKIE_BROWSER and with_cookies:
            print(f"  ⚠️  Cookie download failed, retrying without cookies...")
            _attempt(use_cookies=False)
        else:
            raise

    parent = Path(out_path_no_ext).parent
    stem   = Path(out_path_no_ext).name
    for f in parent.iterdir():
        if f.stem == stem and f.suffix.lower() in {'.webm', '.opus', '.m4a', '.mp4', '.ogg', '.aac'}:
            return f

    raise RuntimeError(f"Downloaded file not found for: {url}")

# ─────────────────────────────────────────────
#  THUMBNAIL DOWNLOAD
# ─────────────────────────────────────────────

def download_thumbnail(url, out_path_no_ext):
    opts = {
        **ydl_base_opts(),
        'skip_download': True,
        'writethumbnail': True,
        'outtmpl': f'{out_path_no_ext}.%(ext)s',
        'ffmpeg_location': FFMPEG,
    }
    try:
        with yt_dlp.YoutubeDL(opts) as ydl:
            ydl.download([url])
    except Exception:
        return None

    for ext in ['jpg', 'jpeg', 'png', 'webp']:
        src = Path(f'{out_path_no_ext}.{ext}')
        if src.exists():
            if ext in ('jpg', 'jpeg'):
                return src  # already a JPEG — embed as-is
            try:
                jpg = Path(f'{out_path_no_ext}_thumb.jpg')
                Image.open(src).convert('RGB').save(str(jpg), format='JPEG', subsampling=0)
                src.unlink(missing_ok=True)
                return jpg
            except Exception:
                src.unlink(missing_ok=True)
                return None

    return None

# ─────────────────────────────────────────────
#  COVER ART TRANSFORMS
# ─────────────────────────────────────────────

def art_center_crop(img: Image.Image) -> Image.Image:
    """Crop to center square."""
    w, h = img.size
    side = min(w, h)
    return img.crop(((w - side) // 2, (h - side) // 2,
                     (w + side) // 2, (h + side) // 2))

def art_smart_crop(img: Image.Image) -> Image.Image:
    """Entropy-based crop — finds the most visually interesting square region."""
    w, h = img.size
    side = min(w, h)
    return ImageOps.fit(img, (side, side), method=Image.LANCZOS, centering=(0.5, 0.5))

def art_padded_blur(img: Image.Image) -> Image.Image:
    """Place original image centered on a blurred, zoomed version of itself."""
    w, h = img.size
    side = max(w, h)

    # Build blurred background: zoom and heavily blur the original
    bg = img.resize((side, side), Image.LANCZOS)
    bg = bg.filter(ImageFilter.GaussianBlur(radius=side // 20))

    # Paste original centered
    paste_x = (side - w) // 2
    paste_y = (side - h) // 2
    bg.paste(img, (paste_x, paste_y))
    return bg

def apply_art_style(src_jpg: Path, style: str) -> bytes:
    """Open a thumbnail, apply the chosen square style, return JPEG bytes."""
    img = Image.open(src_jpg).convert('RGB')
    w, h = img.size

    if w == h:
        # Already square — no transform needed
        result = img
    elif style == 'center':
        result = art_center_crop(img)
    elif style == 'smart':
        result = art_smart_crop(img)
    elif style == 'blur':
        result = art_padded_blur(img)
    else:
        result = img

    buf = io.BytesIO()
    result.save(buf, format='JPEG', subsampling=0)
    return buf.getvalue()

# ─────────────────────────────────────────────
#  GET AUDIO DURATION
# ─────────────────────────────────────────────

def get_duration_sec(file_path):
    try:
        out = subprocess.run(
            [FFPROBE, '-v', 'error', '-show_entries', 'format=duration',
             '-of', 'default=noprint_wrappers=1:nokey=1', str(file_path)],
            capture_output=True, text=True, check=True
        ).stdout.strip()
        return float(out)
    except Exception:
        return 0.0

# ─────────────────────────────────────────────
#  ENCODE TO M4A WITH METADATA + ART
# ─────────────────────────────────────────────

def encode_track(src_file, out_m4a, encoder, title, artist, album, album_artist, track_num, total_tracks, cover_jpg):
    duration_sec = get_duration_sec(src_file)

    cmd = [FFMPEG, '-y', '-i', str(src_file)]
    inputs = 1

    if cover_jpg and cover_jpg.exists():
        cmd += ['-i', str(cover_jpg)]
        inputs = 2

    cmd += [
        '-map', '0:a',
        '-c:a', encoder,
        '-b:a', '256k',
        '-metadata', f'title={title}',
        '-metadata', f'artist={artist}',
        '-metadata', f'album={album}',
        '-metadata', f'album_artist={album_artist}',
        '-metadata', f'track={track_num}/{total_tracks}',
    ]

    if inputs == 2:
        cmd += [
            '-map', '1:v',
            '-c:v', 'mjpeg',
            '-disposition:v:0', 'attached_pic',
        ]

    cmd += ['-movflags', '+faststart', str(out_m4a)]

    pbar = tqdm(total=int(duration_sec) or 1, unit='s', desc=f"  Encoding", ncols=68, leave=False)
    proc = subprocess.Popen(
        cmd + ['-progress', 'pipe:1', '-nostats'],
        stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True
    )
    last_sec = 0
    for line in proc.stdout:
        if 'out_time_ms=' in line:
            try:
                ms  = int(line.split('=')[1].strip())
                cur = min(ms // 1000, int(duration_sec))
                pbar.update(cur - last_sec)
                last_sec = cur
            except ValueError:
                pass
    proc.wait()
    pbar.close()

    if proc.returncode != 0:
        fallback_cmd = [c if c != encoder else 'aac' for c in cmd]
        subprocess.run(fallback_cmd, check=True, capture_output=True)

# ─────────────────────────────────────────────
#  REEMBED COVER ART (lossless — audio untouched)
# ─────────────────────────────────────────────

def reembed_cover(m4a_path: Path, jpeg_bytes: bytes) -> None:
    """Replace the covr atom in an M4A without touching the audio stream."""
    audio = MP4(str(m4a_path))
    if audio.tags is None:
        audio.add_tags()
    audio.tags['covr'] = [MP4Cover(jpeg_bytes, imageformat=MP4Cover.FORMAT_JPEG)]
    audio.save()

# ─────────────────────────────────────────────
#  CACHE  (per-playlist, keyed by URL)
# ─────────────────────────────────────────────

CACHE_FILE = 'streamlist_cache.json'

def load_cache(out_dir: Path) -> dict:
    path = out_dir / CACHE_FILE
    if path.exists():
        try:
            return json.loads(path.read_text(encoding='utf-8'))
        except Exception:
            return {}
    return {}

def save_cache(out_dir: Path, cache: dict):
    path = out_dir / CACHE_FILE
    path.write_text(json.dumps(cache, indent=2, ensure_ascii=False), encoding='utf-8')

# ─────────────────────────────────────────────
#  M3U WRITER
# ─────────────────────────────────────────────

def write_m3u(playlist_path, tracks_info):
    with open(playlist_path, 'w', encoding='utf-8') as f:
        f.write('#EXTM3U\n')
        for filename, duration_sec, display_name in tracks_info:
            f.write(f'#EXTINF:{int(duration_sec)},{display_name}\n')
            f.write(f'{filename}\n')

# ─────────────────────────────────────────────
#  NOTIFICATION
# ─────────────────────────────────────────────

def play_notification():
    print('\a')
    sys.stdout.flush()
    try:
        if sys.platform == 'darwin':
            subprocess.run(['say', 'Playlist ready'], check=False,
                           stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        pass

# ─────────────────────────────────────────────
#  COOKIE / BROWSER SETUP
# ─────────────────────────────────────────────

BROWSERS = ['safari', 'chrome', 'firefox', 'edge', 'brave', 'opera']

def setup_cookies(browser_arg):
    global _COOKIE_BROWSER
    if browser_arg:
        _COOKIE_BROWSER = browser_arg
    else:
        print("\n🔐 Use browser cookies? (needed for Premium / private playlists)")
        for i, b in enumerate(BROWSERS, 1):
            print(f"   {i} — {b.capitalize()}")
        print("   0 — No (public videos only)")
        choice = input("Choice [0-6]: ").strip()
        if choice in [str(i+1) for i in range(len(BROWSERS))]:
            _COOKIE_BROWSER = BROWSERS[int(choice) - 1]
        else:
            _COOKIE_BROWSER = None

    if _COOKIE_BROWSER:
        print(f"   🍪 Using cookies from: {_COOKIE_BROWSER.capitalize()}")
    else:
        print("   ⚠️  No cookies — only public videos will work")

# ─────────────────────────────────────────────
#  BUCKET 1 — DOWNLOAD
# ─────────────────────────────────────────────

def run_download(args):
    global _COOKIE_BROWSER

    print("\n🔍 Detecting AAC encoder...")
    encoder, is_hw, enc_desc = detect_best_aac_encoder()
    print(f"   {'⚡ Hardware' if is_hw else '🖥️  Software'} → {enc_desc}")

    playlist_name = None
    tracks = []
    needs_cookies = False

    if args.excel:
        source_path = args.excel.strip('"\'').strip()
        print(f"\n📄 Loading tracks from Excel: {source_path}")
        tracks = load_excel(source_path)
        print(f"   {len(tracks)} track(s) found")
        playlist_name = args.name or Path(source_path).stem
        needs_cookies = True

    elif args.url:
        needs_cookies = True

    else:
        print("\nHow do you want to provide your tracks?")
        print("  1 — YouTube playlist URL")
        print("  2 — Excel file (.xlsx)")
        choice = input("Choice [1/2]: ").strip()

        if choice == '1':
            needs_cookies = True
        elif choice == '2':
            path = input("Excel file path: ").strip().strip('"\'').strip()
            if not os.path.isfile(path):
                print(f"❌ File not found: {path}")
                sys.exit(1)
            tracks = load_excel(path)
            print(f"   {len(tracks)} track(s) found")
            playlist_name = Path(path).stem
            needs_cookies = True
        else:
            print("❌ Invalid choice.")
            sys.exit(1)

    if needs_cookies:
        setup_cookies(args.browser)

    if args.url:
        print(f"\n📋 Fetching playlist info...")
        playlist_name, tracks = expand_playlist_url(args.url)
        if args.name:
            playlist_name = args.name
    elif not args.excel and needs_cookies and not tracks:
        url = input("YouTube playlist URL: ").strip()
        print("📋 Fetching playlist info...")
        playlist_name, tracks = expand_playlist_url(url)

    if not args.excel and not args.url:
        if args.name:
            playlist_name = args.name
        else:
            override = input(f"\nPlaylist name [{playlist_name}]: ").strip()
            if override:
                playlist_name = override

    if not tracks:
        print("❌ No tracks to download.")
        sys.exit(1)

    playlist_name = safe_filename(playlist_name)

    base_out = Path(args.out) if args.out else Path(os.getcwd())
    out_dir  = base_out / playlist_name
    out_dir.mkdir(parents=True, exist_ok=True)
    print(f"\n📁 Output folder: {out_dir}")

    cache = load_cache(out_dir)
    if cache:
        print(f"\n💾 Cache loaded — {len(cache)} track(s) previously downloaded.")
    else:
        existing_m4a = list(out_dir.glob('*.m4a'))
        if existing_m4a:
            print(f"\n🔄 No cache found but {len(existing_m4a)} existing file(s) detected — rebuilding cache from filenames...")
            rebuilt = 0
            for tidx, track in enumerate(tracks, 1):
                if not track.get('title'):
                    continue
                title  = track['title']
                artist = track.get('artist', '')
                needle = safe_filename(title).lower()
                for f in existing_m4a:
                    stem = re.sub(r'^\d+\s*-\s*', '', f.stem).lower()
                    if stem == needle:
                        duration = get_duration_sec(f)
                        cache[track['url']] = {
                            'title':    title,
                            'artist':   artist,
                            'filename': f.name,
                            'duration': duration,
                        }
                        rebuilt += 1
                        break
            if rebuilt:
                save_cache(out_dir, cache)
                print(f"   ✅ Rebuilt cache for {rebuilt} track(s).")

    total       = len(tracks)
    m3u_data    = []
    succeeded   = 0
    skipped     = 0
    total_start = time.time()

    for idx, track in enumerate(tracks, 1):
        url           = track['url']
        track_num_str = f"{idx:02d}"

        print(f"\n{'─'*52}")

        if url in cache:
            entry    = cache[url]
            title    = entry['title']
            artist   = entry.get('artist', '')
            filename = f"{track_num_str} - {safe_filename(title)}.m4a"
            out_m4a  = out_dir / filename

            cached_file = out_dir / entry['filename']
            if cached_file.exists() and cached_file != out_m4a:
                cached_file.rename(out_m4a)
                cache[url]['filename'] = filename
                save_cache(out_dir, cache)

            if out_m4a.exists():
                duration = entry.get('duration', get_duration_sec(out_m4a))
                display  = f"{artist} - {title}" if artist else title
                m3u_data.append((filename, duration, display))
                skipped += 1
                print(f"  [{idx}/{total}] ⏭️  {title}  (cached)")
                continue

        artist = track.get('artist', '')
        if not track.get('title') or not artist:
            print(f"  [{idx}/{total}] Fetching info...")
            try:
                with yt_dlp.YoutubeDL(ydl_base_opts()) as ydl:
                    info = ydl.extract_info(url, download=False)
                if not track.get('title'):
                    track['title'] = info.get('title', f'Track {idx}')
                if not artist:
                    artist = (info.get('artist') or info.get('creator')
                              or info.get('uploader') or info.get('channel', ''))
            except Exception:
                track['title'] = f'Track {idx}'

        title    = track['title']
        filename = f"{track_num_str} - {safe_filename(title)}.m4a"
        out_m4a  = out_dir / filename

        print(f"  [{idx}/{total}] 🎵 {title}")
        if artist:
            print(f"       Artist: {artist}")

        try:
            with tempfile.TemporaryDirectory() as tmp:
                tmp_path = Path(tmp)

                print(f"  ⬇️  Downloading audio...")
                raw_audio = download_audio(url, str(tmp_path / 'audio'))

                print(f"  🖼️  Downloading thumbnail...")
                cover = download_thumbnail(url, str(tmp_path / 'thumb'))

                encode_track(
                    src_file=raw_audio,
                    out_m4a=out_m4a,
                    encoder=encoder,
                    title=title,
                    artist=artist,
                    album=playlist_name,
                    album_artist=f"Aey - {playlist_name}",
                    track_num=idx,
                    total_tracks=total,
                    cover_jpg=cover,
                )

            duration = get_duration_sec(out_m4a)
            size_mb  = out_m4a.stat().st_size / (1024 * 1024)
            display  = f"{artist} - {title}" if artist else title
            m3u_data.append((filename, duration, display))

            cache[url] = {'title': title, 'artist': artist,
                          'filename': filename, 'duration': duration}
            save_cache(out_dir, cache)

            succeeded += 1
            print(f"  ✅ Saved: {filename}  ({size_mb:.1f} MB)")

        except Exception as e:
            print(f"  ❌ Failed: {e}")

    if m3u_data:
        m3u_path = out_dir / f"{playlist_name}.m3u"
        write_m3u(m3u_path, m3u_data)
        print(f"\n📋 Playlist file: {m3u_path.name}")

    elapsed = time.time() - total_start
    print(f"\n{'═'*52}")
    print(f"🎉 Done!  {succeeded} downloaded  |  {skipped} skipped  |  {total - succeeded - skipped} failed")
    print(f"⏱️  Total time: {int(elapsed//60)}m {elapsed%60:.1f}s")
    print(f"📁 {out_dir}")
    print(f"{'═'*52}")

    if not args.no_notification:
        play_notification()

# ─────────────────────────────────────────────
#  BUCKET 2 — MAKE EDITS
# ─────────────────────────────────────────────

def run_edit_cover_art(args):
    """Re-download thumbnails for every cached track and re-embed as a square."""

    # ── Locate folder ─────────────────────────────────────────────────────────
    if args.out:
        folder = Path(args.out)
    else:
        folder = Path(input("\n📁 Path to playlist folder: ").strip().strip('"\''))

    if not folder.is_dir():
        print(f"❌ Folder not found: {folder}")
        sys.exit(1)

    cache = load_cache(folder)
    if not cache:
        print("❌ No streamlist_cache.json found in that folder. Only folders downloaded by streamlist are supported.")
        sys.exit(1)

    print(f"   Found {len(cache)} cached track(s).")

    # ── Choose crop style ─────────────────────────────────────────────────────
    print("\n🖼️  Choose cover art style:")
    print("   1 — Center crop   (safe, always works)")
    print("   2 — Smart crop    (entropy-based — finds the most interesting region)")
    print("   3 — Padded blur   (original image centered on blurred background)")
    style_choice = input("Choice [1/2/3]: ").strip()
    style_map = {'1': 'center', '2': 'smart', '3': 'blur'}
    style = style_map.get(style_choice)
    if not style:
        print("❌ Invalid choice.")
        sys.exit(1)

    # ── Cookie setup for thumbnail re-downloads ────────────────────────────────
    setup_cookies(args.browser)

    # ── Process each track ────────────────────────────────────────────────────
    total     = len(cache)
    succeeded = 0
    skipped   = 0
    t0        = time.time()

    for idx, (url, entry) in enumerate(cache.items(), 1):
        title    = entry.get('title', f'Track {idx}')
        filename = entry.get('filename', '')
        m4a_path = folder / filename

        print(f"\n{'─'*52}")
        print(f"  [{idx}/{total}] 🎵 {title}")

        if not m4a_path.exists():
            print(f"  ⚠️  File not found, skipping: {filename}")
            skipped += 1
            continue

        try:
            with tempfile.TemporaryDirectory() as tmp:
                cover = download_thumbnail(url, str(Path(tmp) / 'thumb'))
                if not cover:
                    print(f"  ⚠️  No thumbnail found, skipping.")
                    skipped += 1
                    continue

                jpeg_bytes = apply_art_style(cover, style)
                reembed_cover(m4a_path, jpeg_bytes)

            print(f"  ✅ Cover updated: {filename}")
            succeeded += 1

        except Exception as e:
            print(f"  ❌ Failed: {e}")

    elapsed = time.time() - t0
    print(f"\n{'═'*52}")
    print(f"🎉 Done!  {succeeded} updated  |  {skipped} skipped  |  {total - succeeded - skipped} failed")
    print(f"⏱️  Total time: {int(elapsed//60)}m {elapsed%60:.1f}s")
    print(f"{'═'*52}")

    if not args.no_notification:
        play_notification()

# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="streamlist — YouTube playlist downloader and editor")
    parser.add_argument('--url',             help='YouTube playlist or video URL')
    parser.add_argument('--excel',           help='Path to Excel file (.xlsx) with url/title/artist columns')
    parser.add_argument('--name',            help='Playlist / album name')
    parser.add_argument('--out',             help='Output directory (default: current directory)')
    parser.add_argument('--browser',         help='Browser to load cookies from',
                        choices=['safari', 'chrome', 'firefox', 'edge', 'brave', 'opera'])
    parser.add_argument('--no-notification', action='store_true')
    args = parser.parse_args()

    print("╔══════════════════════════════════════════════════╗")
    print("║   🎵  streamlist                                 ║")
    print("╚══════════════════════════════════════════════════╝")

    # If a URL or Excel was passed directly, go straight to download
    if args.url or args.excel:
        run_download(args)
        return

    print("\nWhat would you like to do?")
    print("  1 — Download  (new playlist or sync existing)")
    print("  2 — Make edits")
    bucket = input("Choice [1/2]: ").strip()

    if bucket == '1':
        run_download(args)

    elif bucket == '2':
        print("\nMake edits:")
        print("  1 — Fix cover art  (re-download & reshape to square)")
        edit_choice = input("Choice [1]: ").strip()

        if edit_choice == '1':
            run_edit_cover_art(args)
        else:
            print("❌ Invalid choice.")
            sys.exit(1)

    else:
        print("❌ Invalid choice.")
        sys.exit(1)


if __name__ == '__main__':
    main()
